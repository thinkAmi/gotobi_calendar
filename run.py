import datetime
import itertools

import openpyxl


def main():
    input_date = input_values()
    if not input_date:
        print('終了します')
        return

    calendar = create_calendar(input_date)
    to_excel(calendar)
    print('作成しました')


def input_values():
    print('開始する年を入力してください')
    yyyy = input()

    print('開始する月を入力してください')
    mm = input()

    print('開始する日を入力してください')
    dd = input()

    try:
        if int(dd) > 28:
            print('29日以降はサポート対象外です')
            return None
        return datetime.datetime(int(yyyy), int(mm), int(dd))
    except:
        print('日付ではありません')
        return None


def create_calendar(current_date):
    start_at = current_date.day
    calendar = []

    for _ in range(12):
        # カレンダーの初日から最終日までの日付を作る
        dates = []
        while True:
            dates.append(current_date)
            current_date += datetime.timedelta(days=1)

            # ここまでで翌月の当日になっていたら、処理を終了する
            if start_at == current_date.day:
                break

        # print(dates)
        # => [
        #     datetime.datetime(2021, 7, 10, 0, 0), ... , datetime.datetime(2021, 8, 9, 0, 0)
        #    ]


        # 先頭を埋める
        # 先頭は日曜日始まり
        weekday_of_first_day = dates[0].weekday()
        if weekday_of_first_day != 6:  # 日曜日以外
            # 日曜日始まりでない場合、開始日の曜日以前はダミー(None)を入れておく
            for _ in range(weekday_of_first_day + 1):
                dates.insert(0, None)

        # print(dates)
        # => [
        #     None, None, None, None, None, None,
        #     datetime.datetime(2021, 7, 10, 0, 0), ... , datetime.datetime(2021, 8, 9, 0, 0)
        #    ]


        # リストを一週間ごと(7要素ごと)のリストへ分割し、最後の要素が足りない場合はNoneを入れる
        # 参考：(URL先はPython2なので動かないのに注意)
        # https://iogi.hatenablog.com/entry/split-list-into-sublist
        dates_by_calendar = [item for item in itertools.zip_longest(*[iter(dates)] * 7)]
        # print(dates_by_calendar)
        # [
        #  (None, None, None, None, None, None, datetime.datetime(2021, 7, 10, 0, 0)),
        #  (datetime.datetime(2021, 7, 11, 0, 0), ... , datetime.datetime(2021, 7, 17, 0, 0)),
        #  ...
        #  (datetime.datetime(2021, 8, 8, 0, 0), datetime.datetime(2021, 8, 9, 0, 0), None, None, None, None, None)
        # ]

        calendar.append(dates_by_calendar)

    # print(calendar)
    # [
    #  [
    #   (None, None, None, None, None, None, datetime.datetime(2021, 7, 10, 0, 0)),
    #   (datetime.datetime(2021, 7, 11, 0, 0), ... , datetime.datetime(2021, 7, 17, 0, 0)),
    #   (datetime.datetime(2021, 7, 18, 0, 0), ... , datetime.datetime(2021, 7, 24, 0, 0)),
    #    ...
    #   (datetime.datetime(2021, 8, 8, 0, 0), ... , None)
    #  ],
    #  [
    #   (None, None, datetime.datetime(2021, 8, 10, 0, 0), ... , datetime.datetime(2021, 8, 14, 0, 0)),
    #   ...
    #   (datetime.datetime(2021, 9, 5, 0, 0), ... , None)
    #  ],
    #  ...
    #  [
    #   (None, ..., datetime.datetime(2022, 6, 11, 0, 0)),
    #   ...
    #   (datetime.datetime(2022, 7, 3, 0, 0), ... , datetime.datetime(2022, 7, 9, 0, 0))
    #  ]
    # ]
    return calendar


def to_excel(calendar):
    wb = openpyxl.load_workbook('template_cal.xlsx')

    sheet = wb.copy_worksheet(wb['テンプレート'])
    sheet.title = '結果'

    plot(sheet, calendar)

    wb.save(f'cal_{datetime.datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx')


def plot(sheet, calendar):
    for i, weeks_of_month in enumerate(calendar, 1):
        is_first_time = True

        mod_col = i % 3
        if mod_col == 1:  # 左端のカレンダーに入力
            pos_col = 2
        elif mod_col == 2:  # 中央のカレンダーに入力
            pos_col = 11
        else:
            pos_col = 20  # 右のカレンダーに入力

        if 1 <= i <= 3:  # 1行目のカレンダーに入力
            pos_row = 4
        elif 4 <= i <= 6:  # 2行目のカレンダーに入力
            pos_row = 13
        elif 7 <= i <= 9:  # 3行目のカレンダーに入力
            pos_row = 22
        else:
            pos_row = 31  # 4行目のカレンダーに入力

        # 一ヶ月のうちの一週間分の日付を取得する
        for row_index, current_week in enumerate(weeks_of_month):
            # 日付をセルに設定する
            for col_index, current_date in enumerate(current_week):
                if current_date:  # ダミーは印字しない
                    # そのカレンダーに初めて日付を設定する場合、タイトルも設定する
                    if is_first_time:
                        sheet.cell(row=pos_row-2, column=pos_col, value=f'{current_date.month}月')
                        is_first_time = False

                    sheet.cell(
                        row=pos_row+row_index, column=pos_col+col_index, value=current_date.day
                    )


if __name__ == '__main__':
    main()
